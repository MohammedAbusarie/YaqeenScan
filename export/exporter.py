"""Data export module for attendance records (CSV and XLSX)."""

import csv
import logging
import re
from pathlib import Path

from core.exceptions import YaqeenError
from core.models import AttendanceRecord

logger = logging.getLogger(__name__)


class ExportError(YaqeenError):
    """Raised when export to CSV or XLSX fails."""

    pass


CSV_HEADERS: tuple[str, ...] = (
    "id",
    "session_id",
    "student_id",
    "student_name",
    "mac_address",
    "fingerprint_hash",
    "token_used",
    "submitted_at",
    "ip_address",
)


def export_csv(records: list[AttendanceRecord], filepath: str) -> None:
    """
    Write attendance records to a CSV file.

    Args:
        records: List of attendance records to export.
        filepath: Absolute or relative path for the output CSV file.

    Raises:
        ExportError: If the file cannot be written (e.g. permission or disk error).
    """
    path = Path(filepath)
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        logger.exception("Failed to create export directory: %s", path.parent)
        raise ExportError(f"Cannot create export directory: {e}") from e
    try:
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(CSV_HEADERS)
            for r in records:
                writer.writerow(
                    [
                        r.id,
                        r.session_id,
                        r.student_id,
                        r.student_name,
                        r.mac_address,
                        r.fingerprint_hash,
                        r.token_used,
                        r.submitted_at,
                        r.ip_address,
                    ]
                )
    except OSError as e:
        logger.exception("Failed to write CSV: %s", filepath)
        raise ExportError(f"Cannot write CSV file: {e}") from e


def export_xlsx(records: list[AttendanceRecord], filepath: str) -> None:
    """
    Write attendance records to an XLSX file with header row and light styling.

    Uses openpyxl. Adds a header row, data rows, auto-sized columns, and
    simple header styling (bold font, thin border).

    Args:
        records: List of attendance records to export.
        filepath: Absolute or relative path for the output XLSX file.

    Raises:
        ExportError: If the file cannot be written.
    """
    import openpyxl
    from openpyxl.styles import Border, Font, Side
    from openpyxl.utils import get_column_letter

    path = Path(filepath)
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        logger.exception("Failed to create export directory: %s", path.parent)
        raise ExportError(f"Cannot create export directory: {e}") from e
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws.title == "Sheet":
            ws.title = "Attendance"
        thin = Side(style="thin", color="2A2A2E")
        header_border = Border(
            left=thin, right=thin, top=thin, bottom=thin
        )
        header_font = Font(bold=True)
        for col, header in enumerate(CSV_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.border = header_border
        for row_idx, r in enumerate(records, start=2):
            ws.cell(row=row_idx, column=1, value=r.id)
            ws.cell(row=row_idx, column=2, value=r.session_id)
            ws.cell(row=row_idx, column=3, value=r.student_id)
            ws.cell(row=row_idx, column=4, value=r.student_name)
            ws.cell(row=row_idx, column=5, value=r.mac_address)
            ws.cell(row=row_idx, column=6, value=r.fingerprint_hash)
            ws.cell(row=row_idx, column=7, value=r.token_used)
            ws.cell(row=row_idx, column=8, value=r.submitted_at)
            ws.cell(row=row_idx, column=9, value=r.ip_address)
        row_tuples = [
            (r.id, r.session_id, r.student_id, r.student_name, r.mac_address,
             r.fingerprint_hash, r.token_used, r.submitted_at, r.ip_address)
            for r in records
        ]
        for col in range(1, len(CSV_HEADERS) + 1):
            col_letter = get_column_letter(col)
            max_len = len(str(CSV_HEADERS[col - 1]))
            for row in row_tuples:
                val = row[col - 1]
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[col_letter].width = min(max_len + 1, 50)
        wb.save(path)
    except OSError as e:
        logger.exception("Failed to write XLSX: %s", filepath)
        raise ExportError(f"Cannot write XLSX file: {e}") from e
    except Exception as e:
        logger.exception("Unexpected error writing XLSX: %s", filepath)
        raise ExportError(f"Export failed: {e}") from e


def _safe_course_basename(course_name: str) -> str:
    """Replace non-alphanumeric characters with underscores for safe filenames."""
    safe = re.sub(r"[^\w\s\-]", "", course_name)
    safe = re.sub(r"[\s\-]+", "_", safe).strip("_") or "attendance"
    return safe[:80]


def generate_export_filename(
    course_name: str, date: str, extension: str
) -> str:
    """
    Build a standardized export filename from course name, date, and extension.

    Course name is sanitized for filesystem use (non-alphanumeric replaced).
    Extension may be with or without leading dot; it is normalized to one dot.

    Args:
        course_name: Display name of the course (will be sanitized).
        date: Date string (e.g. YYYY-MM-DD) for the export.
        extension: File extension, e.g. "csv" or ".xlsx".

    Returns:
        A filename string, e.g. "Course_Name_2025-03-05.csv".
    """
    base = _safe_course_basename(course_name)
    ext = extension.strip().lower()
    if ext and not ext.startswith("."):
        ext = "." + ext
    date_safe = re.sub(r"[^\d\-]", "", date)[:10] or "export"
    return f"{base}_{date_safe}{ext}"
