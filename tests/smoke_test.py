"""Phase 8 smoke tests for YaqeenScan.

Verifies that every module imports cleanly and that all core subsystems
function correctly in isolation. Run from the project root:

    python tests/smoke_test.py

Each test prints [PASS] or [FAIL] with a brief reason.
No GUI, no network, no OS-level calls are made.
"""

import os
import sys
import tempfile
import threading
import time
import traceback

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


_results: list[tuple[str, bool, str]] = []


def _run(name: str, fn) -> None:
    """Execute fn() and record pass/fail."""
    try:
        fn()
        _results.append((name, True, ""))
        print(f"  [PASS] {name}")
    except Exception as exc:
        detail = f"{type(exc).__name__}: {exc}"
        _results.append((name, False, detail))
        print(f"  [FAIL] {name}")
        print(f"         {detail}")
        if os.environ.get("SMOKE_VERBOSE"):
            traceback.print_exc()


def _section(title: str) -> None:
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}")


def _assert(condition: bool, msg: str = "") -> None:
    if not condition:
        raise AssertionError(msg or "Assertion failed")


# ---------------------------------------------------------------------------
# 1. Import checks
# ---------------------------------------------------------------------------

def _test_import_core_exceptions() -> None:
    from core.exceptions import YaqeenError
    _assert(issubclass(YaqeenError, Exception))


def _test_import_core_config() -> None:
    from core import config
    _assert(hasattr(config, "DB_PATH"))
    _assert(hasattr(config, "TOKEN_ROTATION_SECONDS"))
    _assert(hasattr(config, "SERVER_HOST"))
    _assert(hasattr(config, "SERVER_PORT"))
    _assert(hasattr(config, "HOTSPOT_SSID"))
    _assert(hasattr(config, "THEME_ACCENT"))


def _test_import_core_models() -> None:
    from core.models import Session, AttendanceRecord, BlockedDevice
    s = Session()
    _assert(hasattr(s, "id"))
    a = AttendanceRecord()
    _assert(hasattr(a, "student_id"))
    b = BlockedDevice()
    _assert(hasattr(b, "mac_address"))


def _test_import_core_database() -> None:
    from core.database import (
        init_db, create_session, end_session, add_attendance,
        is_device_blocked, block_device, get_session_attendance,
        is_student_id_submitted, get_active_session,
        DatabaseError, RecordNotFoundError,
    )
    from core.exceptions import YaqeenError
    _assert(issubclass(DatabaseError, YaqeenError))
    _assert(issubclass(RecordNotFoundError, YaqeenError))


def _test_import_core_token_manager() -> None:
    from core.token_manager import TokenManager, TokenError
    from core.exceptions import YaqeenError
    _assert(issubclass(TokenError, YaqeenError))


def _test_import_core_fingerprint() -> None:
    from core.fingerprint import compute_fingerprint_hash
    _assert(callable(compute_fingerprint_hash))


def _test_import_network_arp_scanner() -> None:
    from network.arp_scanner import get_mac_for_ip, ARPScanError
    from core.exceptions import YaqeenError
    _assert(issubclass(ARPScanError, YaqeenError))


def _test_import_network_hotspot() -> None:
    from network.hotspot import HotspotManager, HotspotError, HotspotStartError, HotspotStopError
    from core.exceptions import YaqeenError
    _assert(issubclass(HotspotError, YaqeenError))
    _assert(issubclass(HotspotStartError, HotspotError))
    _assert(issubclass(HotspotStopError, HotspotError))


def _test_import_server_app() -> None:
    from server.app import create_app
    _assert(callable(create_app))


def _test_import_server_routes() -> None:
    from server.routes import bp, SubmissionError, TokenExpiredError
    from core.exceptions import YaqeenError
    _assert(issubclass(SubmissionError, YaqeenError))
    _assert(issubclass(TokenExpiredError, YaqeenError))


def _test_import_export_exporter() -> None:
    from export.exporter import (
        export_csv, export_xlsx, generate_export_filename, ExportError, CSV_HEADERS
    )
    from core.exceptions import YaqeenError
    _assert(issubclass(ExportError, YaqeenError))
    _assert(len(CSV_HEADERS) == 9)


# ---------------------------------------------------------------------------
# 2. Database tests (in-memory SQLite)
# ---------------------------------------------------------------------------

def _make_in_memory_db():
    from core.database import init_db
    return init_db(":memory:", check_same_thread=True)


def _test_db_init() -> None:
    conn = _make_in_memory_db()
    cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = {row[0] for row in cursor.fetchall()}
    _assert("sessions" in tables, "sessions table missing")
    _assert("attendance" in tables, "attendance table missing")
    _assert("blocked_devices" in tables, "blocked_devices table missing")
    conn.close()


def _test_db_create_session() -> None:
    from core.database import create_session
    conn = _make_in_memory_db()
    s = create_session(conn, "CS101")
    _assert(s.id > 0, f"Expected id > 0, got {s.id}")
    _assert(s.course_name == "CS101")
    _assert(s.is_active is True)
    _assert(s.end_time == "")
    conn.close()


def _test_db_end_session() -> None:
    from core.database import create_session, end_session, get_active_session
    conn = _make_in_memory_db()
    s = create_session(conn, "CS101")
    end_session(conn, s.id)
    active = get_active_session(conn)
    _assert(active is None, "Expected no active session after end_session")
    conn.close()


def _test_db_add_and_get_attendance() -> None:
    from core.database import create_session, add_attendance, get_session_attendance
    from core.models import AttendanceRecord
    from datetime import datetime
    conn = _make_in_memory_db()
    s = create_session(conn, "CS101")
    rec = AttendanceRecord(
        id=0,
        session_id=s.id,
        student_id="S12345",
        student_name="Alice Smith",
        mac_address="aa:bb:cc:dd:ee:ff",
        fingerprint_hash="abc123",
        token_used="deadbeef",
        submitted_at=datetime.now().isoformat(),
        ip_address="192.168.137.5",
    )
    add_attendance(conn, rec)
    records = get_session_attendance(conn, s.id)
    _assert(len(records) == 1, f"Expected 1 record, got {len(records)}")
    _assert(records[0].student_id == "S12345")
    _assert(records[0].student_name == "Alice Smith")
    conn.close()


def _test_db_is_student_id_submitted() -> None:
    from core.database import create_session, add_attendance, is_student_id_submitted
    from core.models import AttendanceRecord
    from datetime import datetime
    conn = _make_in_memory_db()
    s = create_session(conn, "CS101")
    _assert(not is_student_id_submitted(conn, s.id, "S99999"))
    rec = AttendanceRecord(
        id=0, session_id=s.id, student_id="S99999", student_name="Bob Jones",
        mac_address="11:22:33:44:55:66", fingerprint_hash="fp1",
        token_used="tok1", submitted_at=datetime.now().isoformat(), ip_address="10.0.0.1",
    )
    add_attendance(conn, rec)
    _assert(is_student_id_submitted(conn, s.id, "S99999"))
    conn.close()


def _test_db_block_and_check_device() -> None:
    from core.database import create_session, block_device, is_device_blocked
    conn = _make_in_memory_db()
    s = create_session(conn, "CS101")
    mac = "aa:bb:cc:dd:ee:ff"
    fp = "somehash"
    _assert(not is_device_blocked(conn, s.id, mac, fp))
    block_device(conn, mac, fp, s.id)
    _assert(is_device_blocked(conn, s.id, mac, fp))
    conn.close()


def _test_db_block_device_idempotent() -> None:
    from core.database import create_session, block_device, is_device_blocked
    conn = _make_in_memory_db()
    s = create_session(conn, "CS101")
    block_device(conn, "aa:bb:cc:dd:ee:ff", "fp", s.id)
    block_device(conn, "aa:bb:cc:dd:ee:ff", "fp", s.id)
    _assert(is_device_blocked(conn, s.id, "aa:bb:cc:dd:ee:ff", "fp"))
    conn.close()


def _test_db_get_active_session_none() -> None:
    from core.database import get_active_session
    conn = _make_in_memory_db()
    result = get_active_session(conn)
    _assert(result is None, f"Expected None, got {result}")
    conn.close()


# ---------------------------------------------------------------------------
# 3. Token manager tests
# ---------------------------------------------------------------------------

def _test_token_initial_token_set() -> None:
    from core.token_manager import TokenManager
    tm = TokenManager()
    token = tm.get_current()
    _assert(len(token) == 32, f"Expected 32-char hex, got len={len(token)}")
    tm.stop()


def _test_token_validate_current() -> None:
    from core.token_manager import TokenManager
    tm = TokenManager()
    current = tm.get_current()
    _assert(tm.validate(current), "Current token should validate")
    tm.stop()


def _test_token_validate_previous_after_rotate() -> None:
    from core.token_manager import TokenManager
    tm = TokenManager()
    old = tm.get_current()
    tm.rotate()
    _assert(tm.validate(old), "Previous token should still validate (grace period)")
    tm.stop()


def _test_token_reject_expired() -> None:
    from core.token_manager import TokenManager
    tm = TokenManager()
    tm.rotate()
    tm.rotate()
    stale = "0000000000000000000000000000000000000000"
    _assert(not tm.validate(stale), "Stale token should not validate")
    tm.stop()


def _test_token_reject_empty() -> None:
    from core.token_manager import TokenManager
    tm = TokenManager()
    _assert(not tm.validate(""), "Empty string should not validate")
    _assert(not tm.validate(None), "None should not validate")
    tm.stop()


def _test_token_auto_rotation() -> None:
    from core.token_manager import TokenManager
    tm = TokenManager()
    before = tm.get_current()
    tm.start_auto_rotation(interval=1)
    time.sleep(1.5)
    after = tm.get_current()
    tm.stop()
    _assert(before != after, "Token should have rotated after 1s interval")


def _test_token_uniqueness() -> None:
    from core.token_manager import TokenManager
    tm = TokenManager()
    tokens = set()
    for _ in range(10):
        tm.rotate()
        tokens.add(tm.get_current())
    tm.stop()
    _assert(len(tokens) == 10, f"Expected 10 unique tokens, got {len(tokens)}")


# ---------------------------------------------------------------------------
# 4. Fingerprint tests
# ---------------------------------------------------------------------------

def _test_fingerprint_determinism() -> None:
    from core.fingerprint import compute_fingerprint_hash
    h1 = compute_fingerprint_hash("Mozilla/5.0", "en-US", "1920x1080")
    h2 = compute_fingerprint_hash("Mozilla/5.0", "en-US", "1920x1080")
    _assert(h1 == h2, "Same inputs must produce same hash")


def _test_fingerprint_length() -> None:
    from core.fingerprint import compute_fingerprint_hash
    h = compute_fingerprint_hash("ua", "lang", "screen")
    _assert(len(h) == 64, f"SHA-256 hex must be 64 chars, got {len(h)}")


def _test_fingerprint_sensitivity() -> None:
    from core.fingerprint import compute_fingerprint_hash
    h1 = compute_fingerprint_hash("Mozilla/5.0", "en-US", "1920x1080")
    h2 = compute_fingerprint_hash("Mozilla/5.0", "en-US", "1280x720")
    _assert(h1 != h2, "Different screen data must produce different hash")


def _test_fingerprint_empty_inputs() -> None:
    from core.fingerprint import compute_fingerprint_hash
    h = compute_fingerprint_hash("", "", "")
    _assert(len(h) == 64, "Empty inputs must still produce valid SHA-256 hash")


def _test_fingerprint_none_inputs() -> None:
    from core.fingerprint import compute_fingerprint_hash
    h = compute_fingerprint_hash(None, None, None)
    _assert(len(h) == 64, "None inputs must still produce valid SHA-256 hash")


# ---------------------------------------------------------------------------
# 5. ARP parser tests (no OS call — unit tests on the parsing logic)
# ---------------------------------------------------------------------------

def _test_arp_normalize_mac_dashes() -> None:
    from network.arp_scanner import _normalize_mac
    result = _normalize_mac("AA-BB-CC-DD-EE-FF")
    _assert(result == "aa:bb:cc:dd:ee:ff", f"Got {result!r}")


def _test_arp_normalize_mac_colons() -> None:
    from network.arp_scanner import _normalize_mac
    result = _normalize_mac("aa:bb:cc:dd:ee:ff")
    _assert(result == "aa:bb:cc:dd:ee:ff", f"Got {result!r}")


def _test_arp_normalize_mac_uppercase() -> None:
    from network.arp_scanner import _normalize_mac
    result = _normalize_mac("AA:BB:CC:DD:EE:FF")
    _assert(result == "aa:bb:cc:dd:ee:ff", f"Got {result!r}")


def _test_arp_parse_sample_output() -> None:
    """Simulate parsing without running arp -a by monkey-patching subprocess."""
    import subprocess
    from unittest.mock import patch, MagicMock
    from network.arp_scanner import get_mac_for_ip

    sample_arp_output = (
        "\nInterface: 192.168.137.1 --- 0xc\n"
        "  Internet Address      Physical Address      Type\n"
        "  192.168.137.5         aa-bb-cc-dd-ee-ff     dynamic\n"
        "  192.168.137.10        11-22-33-44-55-66     dynamic\n"
    )
    mock_result = MagicMock()
    mock_result.returncode = 0
    mock_result.stdout = sample_arp_output
    mock_result.stderr = ""

    with patch("subprocess.run", return_value=mock_result):
        mac = get_mac_for_ip("192.168.137.5")
    _assert(mac == "aa:bb:cc:dd:ee:ff", f"Expected aa:bb:cc:dd:ee:ff, got {mac!r}")


def _test_arp_parse_ip_not_found() -> None:
    """Return None when IP is absent from ARP table."""
    from unittest.mock import patch, MagicMock
    from network.arp_scanner import get_mac_for_ip

    sample_arp_output = (
        "\nInterface: 192.168.137.1 --- 0xc\n"
        "  192.168.137.5         aa-bb-cc-dd-ee-ff     dynamic\n"
    )
    mock_result = MagicMock()
    mock_result.returncode = 0
    mock_result.stdout = sample_arp_output
    mock_result.stderr = ""

    with patch("subprocess.run", return_value=mock_result):
        mac = get_mac_for_ip("10.0.0.99")
    _assert(mac is None, f"Expected None for missing IP, got {mac!r}")


def _test_arp_raises_on_command_failure() -> None:
    """ARPScanError raised when arp -a returns non-zero."""
    from unittest.mock import patch, MagicMock
    from network.arp_scanner import get_mac_for_ip, ARPScanError

    mock_result = MagicMock()
    mock_result.returncode = 1
    mock_result.stdout = ""
    mock_result.stderr = "Access denied"

    with patch("subprocess.run", return_value=mock_result):
        try:
            get_mac_for_ip("192.168.1.1")
            _assert(False, "Expected ARPScanError")
        except ARPScanError:
            pass


# ---------------------------------------------------------------------------
# 6. Flask test client tests
# ---------------------------------------------------------------------------

def _make_flask_test_app():
    """Create a Flask test app backed by an in-memory DB."""
    import threading
    from core.database import init_db, create_session
    from core.token_manager import TokenManager
    from core import config
    from server.app import create_app

    conn = init_db(":memory:", check_same_thread=False)
    db_lock = threading.Lock()
    create_session(conn, "CS101 Test")
    tm = TokenManager()
    app = create_app(conn, tm, config, db_lock=db_lock)
    app.config["TESTING"] = True
    app.config["SECRET_KEY"] = "test-secret-key-for-smoke-tests"
    app.secret_key = "test-secret-key-for-smoke-tests"
    return app, conn, tm


def _test_flask_index_route() -> None:
    app, conn, tm = _make_flask_test_app()
    with app.test_client() as client:
        resp = client.get("/")
        _assert(resp.status_code == 200, f"GET / returned {resp.status_code}")
    conn.close()
    tm.stop()


def _test_flask_attend_invalid_token() -> None:
    app, conn, tm = _make_flask_test_app()
    with app.test_client() as client:
        resp = client.get("/attend?token=invalidtoken")
        _assert(resp.status_code == 400, f"Expected 400 for bad token, got {resp.status_code}")
        body = resp.data.decode()
        _assert("expired" in body.lower() or "invalid" in body.lower(), "Expected error message in body")
    conn.close()
    tm.stop()


def _test_flask_attend_valid_token() -> None:
    app, conn, tm = _make_flask_test_app()
    with app.test_client() as client:
        token = tm.get_current()
        resp = client.get(f"/attend?token={token}")
        _assert(resp.status_code == 200, f"Expected 200 for valid token, got {resp.status_code}")
        body = resp.data.decode()
        _assert("student" in body.lower() or "form" in body.lower() or "id" in body.lower(),
                "Expected form content in response")
    conn.close()
    tm.stop()


def _test_flask_submit_full_flow() -> None:
    """POST /submit with valid data returns success page and sets cookie."""
    from unittest.mock import patch, MagicMock

    app, conn, tm = _make_flask_test_app()

    mock_result = MagicMock()
    mock_result.returncode = 0
    mock_result.stdout = ""
    mock_result.stderr = ""

    with app.test_client() as client:
        token = tm.get_current()
        with patch("subprocess.run", return_value=mock_result):
            get_resp = client.get(f"/attend?token={token}")
        _assert(get_resp.status_code == 200, f"GET /attend returned {get_resp.status_code}")

        csrf_token = None
        body = get_resp.data.decode()
        import re
        m = re.search(r'name="csrf_token"\s+value="([^"]+)"', body)
        if m:
            csrf_token = m.group(1)
        _assert(csrf_token is not None, "CSRF token not found in form HTML")

        with patch("subprocess.run", return_value=mock_result):
            post_resp = client.post("/submit", data={
                "token": token,
                "csrf_token": csrf_token,
                "student_id": "S20240001",
                "student_name": "Ahmed Hassan",
                "screen_resolution": "1080x1920",
                "timezone": "Africa/Cairo",
                "platform": "Android",
            })
        _assert(post_resp.status_code == 200, f"POST /submit returned {post_resp.status_code}")
        body = post_resp.data.decode()
        _assert("attendance" in body.lower() or "success" in body.lower() or "recorded" in body.lower(),
                "Expected success message in response")

    conn.close()
    tm.stop()


def _test_flask_submit_duplicate_student_id() -> None:
    """Second submission with same student_id but different device returns 409.

    The first client submits successfully (device A). The second client uses a
    distinct User-Agent so its fingerprint differs, bypassing the device-block
    check and reaching the student_id duplicate check which returns 409.
    """
    import re
    from unittest.mock import patch, MagicMock

    app, conn, tm = _make_flask_test_app()

    mock_result = MagicMock()
    mock_result.returncode = 0
    mock_result.stdout = ""
    mock_result.stderr = ""

    def _attend_and_submit(client, student_id: str, student_name: str, ua: str):
        token = tm.get_current()
        with patch("subprocess.run", return_value=mock_result):
            get_resp = client.get(
                f"/attend?token={token}",
                headers={"User-Agent": ua, "Accept-Language": "en-US"},
            )
        if get_resp.status_code != 200:
            return get_resp
        body = get_resp.data.decode()
        m = re.search(r'name="csrf_token"\s+value="([^"]+)"', body)
        csrf = m.group(1) if m else ""
        with patch("subprocess.run", return_value=mock_result):
            return client.post(
                "/submit",
                data={
                    "token": token,
                    "csrf_token": csrf,
                    "student_id": student_id,
                    "student_name": student_name,
                    "screen_resolution": "1080x1920",
                    "timezone": "Africa/Cairo",
                    "platform": "Android",
                },
                headers={"User-Agent": ua, "Accept-Language": "en-US"},
            )

    ua_a = "Mozilla/5.0 (Android 12; Device-A)"
    ua_b = "Mozilla/5.0 (Android 12; Device-B)"

    with app.test_client() as client1:
        r1 = _attend_and_submit(client1, "S20240002", "Sara Ali", ua_a)
        _assert(r1.status_code == 200, f"First submission returned {r1.status_code}")

    with app.test_client() as client2:
        r2 = _attend_and_submit(client2, "S20240002", "Sara Ali", ua_b)
        _assert(r2.status_code == 409,
                f"Duplicate student_id from different device should return 409, got {r2.status_code}")

    conn.close()
    tm.stop()


def _test_flask_submit_invalid_student_id() -> None:
    """Submission with invalid student_id (SQL injection attempt) returns 400."""
    from unittest.mock import patch, MagicMock

    app, conn, tm = _make_flask_test_app()

    mock_result = MagicMock()
    mock_result.returncode = 0
    mock_result.stdout = ""
    mock_result.stderr = ""

    with app.test_client() as client:
        token = tm.get_current()
        with patch("subprocess.run", return_value=mock_result):
            get_resp = client.get(f"/attend?token={token}")
        body = get_resp.data.decode()
        import re
        m = re.search(r'name="csrf_token"\s+value="([^"]+)"', body)
        csrf = m.group(1) if m else ""
        with patch("subprocess.run", return_value=mock_result):
            resp = client.post("/submit", data={
                "token": token,
                "csrf_token": csrf,
                "student_id": "'; DROP TABLE attendance; --",
                "student_name": "Hacker",
                "screen_resolution": "1080x1920",
                "timezone": "Africa/Cairo",
                "platform": "Android",
            })
        _assert(resp.status_code == 400, f"Expected 400 for malicious student_id, got {resp.status_code}")

    conn.close()
    tm.stop()


def _test_flask_submit_missing_csrf() -> None:
    """POST /submit without CSRF token returns 400."""
    from unittest.mock import patch, MagicMock

    app, conn, tm = _make_flask_test_app()

    mock_result = MagicMock()
    mock_result.returncode = 0
    mock_result.stdout = ""
    mock_result.stderr = ""

    with app.test_client() as client:
        token = tm.get_current()
        with patch("subprocess.run", return_value=mock_result):
            resp = client.post("/submit", data={
                "token": token,
                "student_id": "S20240003",
                "student_name": "Omar Khaled",
            })
        _assert(resp.status_code == 400, f"Expected 400 for missing CSRF, got {resp.status_code}")

    conn.close()
    tm.stop()


# ---------------------------------------------------------------------------
# 7. Export tests
# ---------------------------------------------------------------------------

def _make_sample_records():
    from core.models import AttendanceRecord
    from datetime import datetime
    return [
        AttendanceRecord(
            id=1, session_id=1, student_id="S001", student_name="Alice",
            mac_address="aa:bb:cc:dd:ee:01", fingerprint_hash="fp001",
            token_used="tok001", submitted_at=datetime.now().isoformat(),
            ip_address="192.168.137.2",
        ),
        AttendanceRecord(
            id=2, session_id=1, student_id="S002", student_name="Bob",
            mac_address="aa:bb:cc:dd:ee:02", fingerprint_hash="fp002",
            token_used="tok002", submitted_at=datetime.now().isoformat(),
            ip_address="192.168.137.3",
        ),
    ]


def _test_export_csv_creates_file() -> None:
    from export.exporter import export_csv
    records = _make_sample_records()
    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "test_export.csv")
        export_csv(records, path)
        _assert(os.path.exists(path), "CSV file not created")
        with open(path, encoding="utf-8") as f:
            content = f.read()
        _assert("student_id" in content, "CSV header missing")
        _assert("S001" in content, "CSV data row missing")
        _assert("S002" in content, "CSV second row missing")


def _test_export_csv_row_count() -> None:
    import csv as csv_mod
    from export.exporter import export_csv
    records = _make_sample_records()
    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "count_test.csv")
        export_csv(records, path)
        with open(path, encoding="utf-8") as f:
            rows = list(csv_mod.reader(f))
        _assert(len(rows) == 3, f"Expected 3 rows (header + 2 data), got {len(rows)}")


def _test_export_xlsx_creates_file() -> None:
    from export.exporter import export_xlsx
    records = _make_sample_records()
    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "test_export.xlsx")
        export_xlsx(records, path)
        _assert(os.path.exists(path), "XLSX file not created")
        _assert(os.path.getsize(path) > 0, "XLSX file is empty")


def _test_export_xlsx_row_count() -> None:
    import openpyxl
    from export.exporter import export_xlsx
    records = _make_sample_records()
    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "rows_test.xlsx")
        export_xlsx(records, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        _assert(ws.max_row == 3, f"Expected 3 rows (header + 2 data), got {ws.max_row}")
        _assert(ws.cell(row=1, column=3).value == "student_id", "Header cell mismatch")
        _assert(ws.cell(row=2, column=3).value == "S001", "Data cell mismatch")


def _test_export_generate_filename() -> None:
    from export.exporter import generate_export_filename
    name = generate_export_filename("CS 101: Intro", "2026-03-05", "csv")
    _assert(name.endswith(".csv"), f"Expected .csv extension, got {name!r}")
    _assert("2026-03-05" in name, f"Expected date in filename, got {name!r}")
    _assert(";" not in name and "'" not in name, "Filename contains unsafe chars")


def _test_export_generate_filename_xlsx() -> None:
    from export.exporter import generate_export_filename
    name = generate_export_filename("Networks Lab", "2026-03-05", ".xlsx")
    _assert(name.endswith(".xlsx"), f"Expected .xlsx extension, got {name!r}")


def _test_export_empty_records() -> None:
    from export.exporter import export_csv, export_xlsx
    with tempfile.TemporaryDirectory() as tmpdir:
        csv_path = os.path.join(tmpdir, "empty.csv")
        xlsx_path = os.path.join(tmpdir, "empty.xlsx")
        export_csv([], csv_path)
        export_xlsx([], xlsx_path)
        _assert(os.path.exists(csv_path), "Empty CSV not created")
        _assert(os.path.exists(xlsx_path), "Empty XLSX not created")


# ---------------------------------------------------------------------------
# Main runner
# ---------------------------------------------------------------------------

def main() -> None:
    """Run all smoke tests and print a summary."""

    _section("1. Import Checks")
    _run("core.exceptions importable", _test_import_core_exceptions)
    _run("core.config importable with all constants", _test_import_core_config)
    _run("core.models importable with all dataclasses", _test_import_core_models)
    _run("core.database importable with all functions", _test_import_core_database)
    _run("core.token_manager importable", _test_import_core_token_manager)
    _run("core.fingerprint importable", _test_import_core_fingerprint)
    _run("network.arp_scanner importable", _test_import_network_arp_scanner)
    _run("network.hotspot importable", _test_import_network_hotspot)
    _run("server.app importable", _test_import_server_app)
    _run("server.routes importable", _test_import_server_routes)
    _run("export.exporter importable", _test_import_export_exporter)

    _section("2. Database (in-memory SQLite)")
    _run("init_db creates all three tables", _test_db_init)
    _run("create_session returns Session with id>0", _test_db_create_session)
    _run("end_session deactivates session", _test_db_end_session)
    _run("add_attendance + get_session_attendance round-trip", _test_db_add_and_get_attendance)
    _run("is_student_id_submitted returns correct bool", _test_db_is_student_id_submitted)
    _run("block_device + is_device_blocked round-trip", _test_db_block_and_check_device)
    _run("block_device is idempotent (no IntegrityError)", _test_db_block_device_idempotent)
    _run("get_active_session returns None when no sessions", _test_db_get_active_session_none)

    _section("3. Token Manager")
    _run("initial token is 32-char hex string", _test_token_initial_token_set)
    _run("validate() accepts current token", _test_token_validate_current)
    _run("validate() accepts previous token (grace period)", _test_token_validate_previous_after_rotate)
    _run("validate() rejects expired token", _test_token_reject_expired)
    _run("validate() rejects empty/None", _test_token_reject_empty)
    _run("start_auto_rotation() rotates token after interval", _test_token_auto_rotation)
    _run("10 consecutive rotations produce 10 unique tokens", _test_token_uniqueness)

    _section("4. Fingerprint")
    _run("same inputs produce same hash (determinism)", _test_fingerprint_determinism)
    _run("hash is 64-char SHA-256 hex", _test_fingerprint_length)
    _run("different screen data produces different hash", _test_fingerprint_sensitivity)
    _run("empty inputs produce valid hash", _test_fingerprint_empty_inputs)
    _run("None inputs produce valid hash", _test_fingerprint_none_inputs)

    _section("5. ARP Parser (mocked subprocess)")
    _run("_normalize_mac handles dash-separated input", _test_arp_normalize_mac_dashes)
    _run("_normalize_mac handles colon-separated input", _test_arp_normalize_mac_colons)
    _run("_normalize_mac lowercases uppercase input", _test_arp_normalize_mac_uppercase)
    _run("get_mac_for_ip parses sample arp -a output", _test_arp_parse_sample_output)
    _run("get_mac_for_ip returns None for missing IP", _test_arp_parse_ip_not_found)
    _run("get_mac_for_ip raises ARPScanError on command failure", _test_arp_raises_on_command_failure)

    _section("6. Flask Routes (test client)")
    _run("GET / returns 200", _test_flask_index_route)
    _run("GET /attend with invalid token returns 400", _test_flask_attend_invalid_token)
    _run("GET /attend with valid token returns 200 form", _test_flask_attend_valid_token)
    _run("POST /submit full happy-path returns 200 success", _test_flask_submit_full_flow)
    _run("POST /submit duplicate student_id returns 409", _test_flask_submit_duplicate_student_id)
    _run("POST /submit with malicious student_id returns 400", _test_flask_submit_invalid_student_id)
    _run("POST /submit without CSRF token returns 400", _test_flask_submit_missing_csrf)

    _section("7. Export")
    _run("export_csv creates file with data", _test_export_csv_creates_file)
    _run("export_csv writes correct row count", _test_export_csv_row_count)
    _run("export_xlsx creates non-empty file", _test_export_xlsx_creates_file)
    _run("export_xlsx writes correct row count and headers", _test_export_xlsx_row_count)
    _run("generate_export_filename produces safe .csv name", _test_export_generate_filename)
    _run("generate_export_filename produces safe .xlsx name", _test_export_generate_filename_xlsx)
    _run("export_csv and export_xlsx handle empty record list", _test_export_empty_records)

    # Summary
    total = len(_results)
    passed = sum(1 for _, ok, _ in _results if ok)
    failed = total - passed

    print(f"\n{'='*60}")
    print(f"  RESULTS: {passed}/{total} passed", end="")
    if failed:
        print(f"  ({failed} FAILED)")
        print()
        for name, ok, detail in _results:
            if not ok:
                print(f"  [FAIL] {name}")
                print(f"         {detail}")
    else:
        print("  — all tests passed")
    print(f"{'='*60}\n")

    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
